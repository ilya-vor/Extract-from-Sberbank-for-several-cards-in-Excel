import fitz
import os

def extract_text_from_pdf(pdf_path, txt_path):
    try:
        # Открываем PDF-файл для чтения бинарного содержимого
        doc = fitz.open(pdf_path)
        # Создаем объект для записи текста в текстовый файл
        with open(txt_path, 'w', encoding='utf-8') as txt_file:
            # Итерируем по страницам и записываем текст в текстовый файл
            for page_number in range(doc.page_count):
                page = doc[page_number]
                text = page.get_text()
                txt_file.write(text)

        print(f'Текст успешно извлечен из PDF и сохранен в файл: {txt_path}')

    except Exception as e:
        print(f'Произошла ошибка: {str(e)}')

# Задаем пути к файлам
pdf_file_path = 'Выписка по дебетовой карте (на русском).pdf'
txt_file_path = 'Выписка по дебетовой карте (на русском).txt'

# Вызываем функцию для извлечения текста из PDF и записи его в текстовый файл
extract_text_from_pdf(pdf_file_path, txt_file_path)



import openpyxl

# Открытие файла
with open('Выписка по дебетовой карте (на русском).txt', 'r', encoding='utf-8') as file:
    # Чтение содержимого файла построчно в список
    lines = file.readlines()

# Создаем новый Excel файл
workbook = openpyxl.Workbook()

# Выбираем активный лист
sheet = workbook.active

# Записываем текст в разные ячейки
sheet['A1'] = "Дата"
sheet['B1'] = "Время"
sheet['C1'] = "Категория"
sheet['D1'] = "Комментарий"
sheet['E1'] = "Сумма"
itxt = 30
iexcel = 2
while True:
    try:
        if lines[itxt] == "Продолжение на следующей странице\n":
            itxt += 11
        sheet[f"A{iexcel}"] = lines[itxt]
        sheet[f"B{iexcel}"] = lines[itxt+1]
        sheet[f"C{iexcel}"] = lines[itxt+4]
        sheet[f"D{iexcel}"] = lines[itxt+5][0:lines[itxt+5].index('.')]
        if lines[itxt+6][0] == "*":
            itxt += 1
        cleaned_string = ''.join(char for char in lines[itxt+6] if char.isdigit() or char in {'-', '.', '+', ','})
        cleaned_string = cleaned_string.replace(',','.')
        if cleaned_string[0] == '+':
            sheet[f"E{iexcel}"] = float(cleaned_string)
        else:
            sheet[f"E{iexcel}"] = float(cleaned_string) * -1
        itxt += 7
        iexcel += 1
    except Exception as e:
        print(f'Произошла ошибка: {str(e)}')
        break
for b in "ABCD":
    sheet[f"{b}{iexcel}"] = ""
# Сохраняем файл
workbook.save('Выписка по дебетовой карте (на русском).xlsx')

print('Файл успешно создан и сохранен.')



import openpyxl

# Открываем Excel файл
workbook = openpyxl.load_workbook('Выписка по дебетовой карте (на русском).xlsx')

Months = ["Декабрь","Ноябрь","Октябрь",
          "Сентябрь","Август","Июль",
          "Июнь","Май","Апрель",
          "Март","Февраль","Январь"]
Months = Months[::-1]

#Открываем общую историю операций по карте 
sheet = workbook['Sheet']

inews = [2 for i in range (12)]
iold = 2
while sheet[f"A{iold}"].value != None:
    cell_value = sheet[f"A{iold}"].value
    Number_of_mounth = int(str(cell_value)[3:5])
    Name_of_mounth = Months[Number_of_mounth - 1]
    try:
        target_sheet = workbook[Name_of_mounth]
    except:
        target_sheet = workbook.create_sheet(title=Name_of_mounth)
        for col in "ABCDE":
            target_sheet[f"{col}1"].value = sheet[f"{col}1"].value
        
    for col in "ABCDE":
        target_sheet[f"{col}{inews[Number_of_mounth - 1]}"].value = sheet[f"{col}{iold}"].value
    inews[Number_of_mounth - 1] += 1
    iold += 1

# Сохраняем файл
workbook.save('Выписка за год (по месяцам).xlsx')

print('Новый лист успешно создан и данные записаны.')



import openpyxl

# Открываем Excel файл
workbook = openpyxl.load_workbook('Выписка за год (по месяцам).xlsx')

Months = ["Декабрь","Ноябрь","Октябрь",
          "Сентябрь","Август","Июль",
          "Июнь","Май","Апрель",
          "Март","Февраль","Январь"]
Months = Months[::-1]

for month in Months:
    sheet = workbook[month]
    sheet['G1'] = "Категории"
    categories = []
    i = 2
    icat = 2
    while sheet[f"C{i}"].value != None:
        if sheet[f"C{i}"].value not in categories:
            categories.append(sheet[f"C{i}"].value)
            commentaries = []
            while any([sheet[f"G{ii}"].value for ii in range(icat,icat+5)]):
                icat += 1
            icat += 1
            sheet[f"G{icat}"].value  = sheet[f"C{i}"].value 
            iccat = 2
            iname = icat
            while sheet[f"C{iccat}"].value != None:
                if sheet[f"C{iccat}"].value == categories[-1]:
                    if sheet[f"D{iccat}"].value in commentaries:
                        icccat = iname

                        while sheet[f"D{iccat}"].value != sheet[f"G{icccat}"].value:
                            icccat += 1

                            
                        sheet[f"H{icccat}"].value = float(sheet[f"H{icccat}"].value) + float(sheet[f"E{iccat}"].value)
                    else:
                        icat += 1
                        for col in "DE":
                            target_col = chr(ord(col) + 3)
                            sheet[f"{target_col}{icat}"].value = sheet[f"{col}{iccat}"].value
                        commentaries.append(sheet[f"D{iccat}"].value)
                iccat += 1
            if sheet[f"G{iname}"].value not in ["Прочие операции\n","Перевод с карты\n","Перевод на карту\n"]:
                iii = iname + 1
                s = 0
                while sheet[f"G{iii}"].value != None:
                    iii += 1
                    s += float(sheet[f"H{iii-1}"].value)
                sheet[f"G{iii}"] = "Итого:"
                sheet[f"H{iii}"] = s
        i += 1
            
workbook.save('Выписка за год (по месяцам, отсортированная).xlsx')

print('Новый лист успешно создан и данные записаны.')



import openpyxl

# Открываем Excel файл
workbook = openpyxl.load_workbook('Выписка за год (по месяцам, отсортированная).xlsx')

Months = ["Декабрь","Ноябрь","Октябрь",
          "Сентябрь","Август","Июль",
          "Июнь","Май","Апрель",
          "Март","Февраль","Январь"]
Months = Months[::-1]

for month in Months:
    print(month)
    sheet = workbook[month]
    sheet['J1'] = "Бюджет"
    sheet['J3'] = "Категории"
    sheet['K3'] = "Сумма"

    i = 2
    dohod = 0
    while any([sheet[f"G{ii}"].value for ii in range(i,i+3)]):
        if sheet[f"H{i}"].value != None:
            if sheet[f"H{i}"].value > 0:
                dohod += sheet[f"H{i}"].value
        i += 1

    sheet['J3'] = "Доход"
    sheet['K3'] = dohod

    i = 2
    while any([sheet[f"G{ii}"].value for ii in range(i,i+3)]):
        if sheet[f"G{i}"].value == "Прочие операции\n":
            ii = i + 1
            s = 0
            while sheet[f"G{ii}"].value != None:
                if sheet[f"H{ii}"].value < 0:
                    s += sheet[f"H{ii}"].value
                ii += 1
            sheet['J4'] = "Прочие операции"
            sheet['K4'] = s
        if sheet[f"G{i}"].value == "Перевод с карты\n":
            sheet['J5'] = "Перевод с карты"
            sheet['K5'].value = sheet[f"H{i+1}"].value
        i += 1
    
    i = 2
    ib = 6
    while any([sheet[f"G{ii}"].value for ii in range(i,i+3)]):
        if sheet[f"G{i}"].value == "Итого:":
            sheet[f'K{ib}'].value = sheet[f"H{i}"].value
            iname = i
            while sheet[f'G{iname}'].value != None:
                iname -= 1
            sheet[f'J{ib}'].value = sheet[f"G{iname+1}"].value
            ib += 1
        i += 1
    sheet[f"J{ib}"] = "Итого:"
    i = 3
    s = 0
    while sheet[f"K{i}"].value != None:
        s += float(sheet[f"K{i}"].value)
        i += 1
    sheet[f"K{i}"] = s
            
            
workbook.save('Выписка за год РЕЗУЛЬТАТ.xlsx')
os.remove("Выписка за год (по месяцам).xlsx")
os.remove("Выписка за год (по месяцам, отсортированная).xlsx")
os.remove("Выписка по дебетовой карте (на русском).txt")
os.remove("Выписка по дебетовой карте (на русском).xlsx")
print('Новый лист успешно создан и данные записаны.')

