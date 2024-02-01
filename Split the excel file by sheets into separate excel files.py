import openpyxl

def save_each_sheet_as_file(input_file):
    # Загрузка рабочей книги (workbook)
    workbook = openpyxl.load_workbook(input_file)

    # Итерация по всем листам
    for sheet_name in workbook.sheetnames:
        # Выбор текущего листа
        sheet = workbook[sheet_name]

        # Создание новой книги только с текущим листом
        new_workbook = openpyxl.Workbook()
        new_workbook.remove(new_workbook.active)  # Удаляем стандартный лист

        # Копирование данных из оригинального листа в новую книгу
        new_sheet = new_workbook.create_sheet(title=sheet_name)  # Создаем новый лист с тем же именем
        for row in sheet.iter_rows():
            new_row = [cell.value for cell in row]
            new_sheet.append(new_row)


        # Определяем максимальную ширину для каждого столбца
        max_width_per_column = [max(len(str(cell.value)) for cell in col) for col in new_sheet.iter_cols()]

        # Устанавливаем ширину каждого столбца в максимальное значение
        for i, max_width in enumerate(max_width_per_column, 1):
            new_sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_width

        # Сохранение новой книги в отдельный файл
        new_file_name = f"{sheet_name} 2023.xlsx"
        new_workbook.save(new_file_name)

    print("Каждый лист сохранен в отдельный файл.")

# Пример использования
input_excel_file = 'Выписка за год РЕЗУЛЬТАТ.xlsx'
save_each_sheet_as_file(input_excel_file)
